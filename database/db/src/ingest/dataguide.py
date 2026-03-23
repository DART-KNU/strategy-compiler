"""
DataGuide Excel file ingest.

Source: raw/dataguide.xlsx
Sheets:
  bm     - Index daily OHLC (KOSPI, KOSPI200, KOSDAQ, KRX300)
  type1  - Stock daily fields (open, close, volume, market_cap, flags, etc.)
  type2  - Quarterly financial data (income statement, balance sheet, cash flow)

File structure (all 3 sheets share same metadata header layout):
  Rows 0-7: metadata / config rows
  Row 8:    column header (코드, 코드명, ..., date1, date2, ...)
  Row 9+:   data (one row per ticker × item)

For type2 (quarterly financials):
  Row 7:  year labels for data columns (e.g. 2018, 2018, 2018, 2018, 2019, ...)
  Row 8:  header row (standard)
  Col 6 in header row = '분기' (visual separator, skip it)
  Cols 7+ in header row = '1Q', '2Q', '3Q', '4Q', repeating per year

The DataGuide format uses a "wide" pivot:
  rows = (ticker × item_code)
  columns = dates (for price data) or (year, quarter) pairs (for financials)

We normalize to long format for SQLite storage.

Performance note:
  The dataguide.xlsx file is ~412 MB. Reading type1 with openpyxl in streaming
  mode (read_only=True) is significantly faster and more memory-efficient than
  loading the full sheet with pandas. Allow ~10-30 minutes for first build.
  Subsequent builds skip re-ingestion if the file checksum is unchanged.
"""

import datetime
import logging
import sqlite3
from pathlib import Path
from typing import Iterator

import pandas as pd

from src.db import insert_batch, truncate_table, get_row_count
from src.utils.io import clean_date_str

logger = logging.getLogger(__name__)

_NOW = lambda: datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
_HEADER_ROW_IDX = 8   # 0-based index of the header row in all DG sheets
_YEAR_ROW_IDX   = 7   # 0-based index of year labels row (type2 only)
_SKIP_COL_IDX   = 6   # column 6 is the '분기' separator in type2


# ---------------------------------------------------------------------------
# Index daily (bm sheet)
# ---------------------------------------------------------------------------

def ingest_index_daily(
    conn: sqlite3.Connection,
    path: Path,
    sheet_name: str,
    index_code_map: dict,
    item_map: dict,
    batch_size: int = 50000,
) -> int:
    """
    Ingest bm sheet into raw_dg_index_daily.

    index_code_map: {'I.001': 'KOSPI', 'I.101': 'KOSPI200', ...}
    item_map: {'시가지수(포인트)': 'open', ...}
    """
    ingested_at = _NOW()
    logger.info("Reading index daily sheet '%s' ...", sheet_name)

    # bm sheet is small; read with pandas
    df_raw = pd.read_excel(str(path), sheet_name=sheet_name, header=None, engine="openpyxl")

    # Header row
    header = df_raw.iloc[_HEADER_ROW_IDX]

    # Build date column map: col_idx -> date_str
    date_col_map = {}
    meta_col_count = 6  # code, name, type, item_code, item_name, period
    for i, v in enumerate(header):
        if i < meta_col_count:
            continue
        ds = _value_to_date_str(v)
        if ds:
            date_col_map[i] = ds

    truncate_table(conn, "raw_dg_index_daily")
    cols = ["raw_code", "code_name", "index_type", "item_code", "item_name",
            "trade_date", "value", "ingested_at"]

    total = 0
    batch = []

    data_rows = df_raw.iloc[(_HEADER_ROW_IDX + 1):].values

    for row in data_rows:
        raw_code   = _cell_str(row, 0)
        code_name  = _cell_str(row, 1)
        index_type = _cell_str(row, 2)
        item_code  = _cell_str(row, 3)
        item_name  = _cell_str(row, 4)

        if not raw_code:
            continue

        for col_idx, trade_date in date_col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if _is_na(val):
                continue

            batch.append((raw_code, code_name, index_type, item_code, item_name,
                          trade_date, float(val), ingested_at))

            if len(batch) >= batch_size:
                insert_batch(conn, "raw_dg_index_daily", batch, cols)
                total += len(batch)
                batch = []

    if batch:
        insert_batch(conn, "raw_dg_index_daily", batch, cols)
        total += len(batch)

    logger.info("Ingested %d rows -> raw_dg_index_daily", total)
    return total


# ---------------------------------------------------------------------------
# Stock daily (type1 sheet) - streaming via openpyxl
# ---------------------------------------------------------------------------

def ingest_stock_daily(
    conn: sqlite3.Connection,
    path: Path,
    sheet_name: str,
    batch_size: int = 50000,
) -> int:
    """
    Ingest type1 sheet into raw_dg_stock_daily using openpyxl streaming.

    Uses read_only=True to avoid loading the full 412MB file into memory.
    This is the most time-consuming step (~10-30 min depending on hardware).
    """
    from openpyxl import load_workbook

    ingested_at = _NOW()
    logger.info("Opening dataguide.xlsx type1 sheet in streaming mode ...")
    logger.info("This is the largest step and may take 10-30 minutes.")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)

    # Skip metadata rows 0-7
    for _ in range(_HEADER_ROW_IDX):
        next(rows_iter)

    # Row 8: header - extract date columns
    header = next(rows_iter)
    date_col_map = {}  # col_idx -> date_str
    meta_col_count = 6

    for i, v in enumerate(header):
        if i < meta_col_count:
            continue
        ds = _value_to_date_str(v)
        if ds:
            date_col_map[i] = ds

    logger.info("Found %d date columns in type1 sheet", len(date_col_map))

    truncate_table(conn, "raw_dg_stock_daily")
    cols = ["raw_ticker", "corp_name", "security_type", "item_code", "item_name",
            "trade_date", "value", "ingested_at"]

    total = 0
    batch = []
    rows_processed = 0

    for row in rows_iter:
        raw_ticker    = _cell_str_row(row, 0)
        corp_name     = _cell_str_row(row, 1)
        security_type = _cell_str_row(row, 2)
        item_code     = _cell_str_row(row, 3)
        item_name     = _cell_str_row(row, 4)

        if not raw_ticker:
            continue

        rows_processed += 1
        if rows_processed % 5000 == 0:
            logger.info("  ... processed %d ticker-item rows (%d total values so far)",
                        rows_processed, total)

        for col_idx, trade_date in date_col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or (isinstance(val, float) and _isnan(val)):
                continue
            try:
                float_val = float(val)
            except (ValueError, TypeError):
                continue  # skip text/annotation cells

            batch.append((raw_ticker, corp_name, security_type, item_code, item_name,
                          trade_date, float_val, ingested_at))

            if len(batch) >= batch_size:
                insert_batch(conn, "raw_dg_stock_daily", batch, cols)
                total += len(batch)
                batch = []

    if batch:
        insert_batch(conn, "raw_dg_stock_daily", batch, cols)
        total += len(batch)

    wb.close()
    logger.info("Ingested %d rows -> raw_dg_stock_daily (%d ticker-item rows)",
                total, rows_processed)
    return total


# ---------------------------------------------------------------------------
# Quarterly financials (type2 sheet)
# ---------------------------------------------------------------------------

def ingest_financials_quarterly(
    conn: sqlite3.Connection,
    path: Path,
    sheet_name: str,
    batch_size: int = 50000,
) -> int:
    """
    Ingest type2 sheet into raw_dg_financials_quarterly.

    The type2 sheet has a two-row header:
      Row 7 (year_row): blank for meta cols, then year values (2018,2018,...,2026)
      Row 8 (header):   meta col names + quarter labels (1Q,2Q,3Q,4Q,...)
      Col 6 = '분기' separator - skipped
    """
    ingested_at = _NOW()
    logger.info("Reading financials quarterly sheet '%s' ...", sheet_name)

    # Read without header to access both row 7 (year) and row 8 (quarter)
    df_raw = pd.read_excel(
        str(path), sheet_name=sheet_name, header=None, engine="openpyxl"
    )

    # Year row (row 7) and header row (row 8)
    year_row   = df_raw.iloc[_YEAR_ROW_IDX]
    header_row = df_raw.iloc[_HEADER_ROW_IDX]

    # Build (col_idx -> (year_str, quarter_str)) map
    period_col_map = {}
    for i, (yr, qtr) in enumerate(zip(year_row, header_row)):
        if i <= 5:
            continue           # meta cols
        if i == _SKIP_COL_IDX:
            continue           # '분기' separator
        if _is_na(yr) or _is_na(qtr):
            continue
        yr_str  = str(int(float(yr)))  # e.g. '2018'
        qtr_str = str(qtr).strip()     # e.g. '1Q'
        if qtr_str in ("1Q", "2Q", "3Q", "4Q"):
            period_col_map[i] = (yr_str, qtr_str)

    logger.info("Found %d (year, quarter) columns in type2 sheet", len(period_col_map))

    truncate_table(conn, "raw_dg_financials_quarterly")
    cols = ["raw_ticker", "corp_name", "fiscal_month", "report_type",
            "item_code", "item_name", "year", "quarter", "value", "ingested_at"]

    total = 0
    batch = []
    data_df = df_raw.iloc[(_HEADER_ROW_IDX + 1):]

    for _, row in data_df.iterrows():
        raw_ticker   = _cell_str(row.values, 0)
        corp_name    = _cell_str(row.values, 1)
        fiscal_month = _cell_str(row.values, 2)
        report_type  = _cell_str(row.values, 3)
        item_code    = _cell_str(row.values, 4)
        item_name    = _cell_str(row.values, 5)

        if not raw_ticker:
            continue

        for col_idx, (yr_str, qtr_str) in period_col_map.items():
            val = row.iloc[col_idx] if col_idx < len(row) else None
            if _is_na(val):
                continue

            batch.append((
                raw_ticker, corp_name, fiscal_month, report_type,
                item_code, item_name, yr_str, qtr_str, float(val), ingested_at
            ))

            if len(batch) >= batch_size:
                insert_batch(conn, "raw_dg_financials_quarterly", batch, cols)
                total += len(batch)
                batch = []

    if batch:
        insert_batch(conn, "raw_dg_financials_quarterly", batch, cols)
        total += len(batch)

    logger.info("Ingested %d rows -> raw_dg_financials_quarterly", total)
    return total


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _value_to_date_str(v) -> str | None:
    """Convert a cell value (datetime or date or string) to ISO-8601 date string."""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return clean_date_str(v)
    return None


def _cell_str(row_array, idx: int) -> str | None:
    """Extract a string value from a numpy/list row by index."""
    try:
        v = row_array[idx]
    except (IndexError, KeyError):
        return None
    if v is None:
        return None
    import numpy as np
    if isinstance(v, float) and (v != v):  # NaN check
        return None
    s = str(v).strip()
    return s if s else None


def _cell_str_row(row, idx: int) -> str | None:
    """Extract a string from an openpyxl streaming row (tuple)."""
    try:
        v = row[idx]
    except (IndexError, TypeError):
        return None
    if v is None:
        return None
    if isinstance(v, float) and (v != v):
        return None
    s = str(v).strip()
    return s if s else None


def _is_na(v) -> bool:
    """Return True if v is None, NaN, or pandas NA."""
    if v is None:
        return True
    if isinstance(v, float) and (v != v):
        return True
    try:
        import pandas as pd
        return pd.isna(v)
    except (TypeError, ValueError):
        return False


def _isnan(v: float) -> bool:
    return v != v  # NaN != NaN
